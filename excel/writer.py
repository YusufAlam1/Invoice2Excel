from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from core.models import InvoiceRow

HEADERS = [
    "Invoice #",
    "Period",
    "Hours Worked",
    "Rate",
    "Amount",
    "HST",
    "Date & Time Paid",
]

NUMERIC_COLS = {
    "Hours Worked",
    "Rate",
    "Amount",
    "HST",
}

EXCEL_NUM_FORMAT = "#,##0.00"


def row_to_dict(row: InvoiceRow) -> dict[str, object]:
    return {
        "Invoice #": row.invoice_no,
        "Period": row.period,        # keep as text (two dates with a dash)
        "Hours Worked": row.hours,
        "Rate": row.rate,
        "Amount": row.amount,
        "HST": row.hst,
        "Date & Time Paid": "",      # left blank for user
    }


def _ensure_employee_sheet(wb: Workbook, sheet_name: str) -> Worksheet:
    """Create the sheet with the exact header row if it doesn't exist."""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(HEADERS)
        # widths for A..G (7 columns)
        widths = [14, 22, 14, 10, 12, 10, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        return ws
    return wb[sheet_name]


def _invoice_exists(ws: Worksheet, invoice_no: str) -> bool:
    """Check Column A ("Invoice #") for an existing value."""
    if ws.max_row < 2:
        return False
    for cell in ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=1, values_only=True):
        existing = cell[0]
        if existing is not None and str(existing).strip() == str(invoice_no).strip():
            return True
    return False


def _apply_cell_formats(ws: Worksheet, row_idx: int):
    """After writing a row, apply basic number formats to the just-written cells."""
    header_to_col = {ws.cell(row=1, column=c).value: c
                     for c in range(1, ws.max_column + 1)}
    for name in NUMERIC_COLS:
        c = header_to_col.get(name)
        if c:
            ws.cell(row=row_idx, column=c).number_format = EXCEL_NUM_FORMAT


def save_employee_row(excel_path: Path, row: InvoiceRow):
    if not excel_path.exists():
        raise FileNotFoundError(f"No workbook found at: {excel_path}")

    wb = load_workbook(excel_path)
    ws = _ensure_employee_sheet(wb, row.employee)

    if _invoice_exists(ws, row.invoice_no):
        raise ValueError("invoice has already been written to file")

    values = row_to_dict(row)
    next_r = ws.max_row + 1

    for col_idx, header in enumerate(HEADERS, start=1):
        val = values.get(header, "")
        if header in NUMERIC_COLS:
            try:
                val = float(val)
            except Exception:
                pass
            ws.cell(row=next_r, column=col_idx, value=val)
        else:
            ws.cell(row=next_r, column=col_idx, value=val)

    _apply_cell_formats(ws, next_r)
    wb.save(excel_path)
    wb.close()